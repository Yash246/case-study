"""
XLSX ingestion class.
"""

import logging
import polars as pl
from pathlib import Path
from typing import Any
from openpyxl import load_workbook

from .data_ingestion import DataIngestion

logger = logging.getLogger(__name__)


class XLSXIngestion(DataIngestion):
    """
    Class for ingestion of accounts, orders and sales history information.

    Args:
        raw_folder: Folder containing raw XLSX files.
        bronze_folder: Folder for Bronze-layer Parquet output.
    """

    def __init__(self, raw_folder: Path, bronze_folder: Path):

        super().__init__(raw_folder, bronze_folder)
        self._sheet_results: dict[str, pl.DataFrame] = {}

    def _read_source(self, path: Path, **kwargs: Any) -> dict[str, pl.DataFrame]:
        """
        Read requested sheets from an Excel workbook.

        Opens the workbook in read-only mode, iterates over the
        requested sheet names, and converts each sheet's rows into
        a Polars DataFrame. Column headers are normalised to
        lowercase with underscores.

        Args:
            path: Absolute path to the Excel workbook.
            **kwargs: Must include sheets (list[str]) with the
                names of sheets to extract.

        Returns:
            Mapping of sheet name to parsed DataFrame.
        """

        sheets: list[str] = kwargs.get("sheets", [])
        results: dict[str, pl.DataFrame] = {}

        wb = load_workbook(path, read_only=True, data_only=True)
        logger.debug(f"Available sheets in workbook: {wb.sheetnames}")

        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet {sheet_name} not found in workbook — skipping ")
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            # not sufficient data
            if len(rows) < 2:
                continue
            
            # clean and standardize
            headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0]]

            # Convert every cell value to string to avoid mixed type schema errors.
            data_rows = []
            for row in rows[1:]:
                data_rows.append(
                    {
                        h: (str(v).strip() if v is not None else None)
                        for h, v in zip(headers, row)
                    }
                )

            df = pl.DataFrame(data_rows, infer_schema_length=0)

            logger.info(f"Sheet {sheet_name} parsed")

            results[sheet_name] = df

        wb.close()
        self._sheet_results = results

        return results

    def _write_bronze(self, data: dict[str, pl.DataFrame]) -> None:
        """
        Write each sheet dataframe to Bronze layer as a separate parquet file.

        Args:
            data: Mapping of sheet name to dataframe.
        """

        for sheet_name, df in data.items():
            output_path = self._bronze_folder / f"{sheet_name}.parquet"
            df.write_parquet(output_path)

    def validate_source(self, path: Path) -> None:
        """
        Check that the XLSX file exists

        Args:
            path: path to the XLSX file.
        """
        if not path.exists():
            logger.warning(f"XLSX file not found: {path}")

    def ingest(self, filename: str, **kwargs: Any) -> dict[str, pl.DataFrame]:
        """
        Ingest an Excel workbook by extracting specified sheets.

        Args:
            filename: Name of the XLSX file within raw_folder.
            **kwargs: Must include sheets (list[str]) with
                sheet names to extract from the workbook.

        Returns:
            Mapping of sheet name to Bronze DataFrame.
        """

        sheets = kwargs.get("sheets", [])

        source_path = self._resolve_path(filename)
        self.validate_source(source_path)

        data = self._read_source(source_path, sheets=sheets)

        self._write_bronze(data)

        logger.info("XLSX ingestion complete")

        return data
